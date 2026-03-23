"""
Fantasy points API endpoints.

Handles fantasy points management, awards, and data fetching.
Business logic is delegated to FantasyService.
"""

import io

from flask import Response, jsonify, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.logger import get_logger
from app.models import FantasyAward, Team
from app.routes import api_bp
from app.routes.main import get_current_league
from app.services.fantasy_service import fantasy_service
from app.utils import admin_required, error_response, is_admin

logger = get_logger(__name__)

# Characters that trigger formula evaluation in Excel
_FORMULA_PREFIXES = ('=', '+', '-', '@', '\t', '\r')


def _sanitize_excel_value(value: str) -> str:
    """Sanitize a string value to prevent Excel formula injection."""
    if isinstance(value, str) and value and value[0] in _FORMULA_PREFIXES:
        return "'" + value
    return value


# ==================== FANTASY POINTS CRUD ====================

@api_bp.route('/fantasy/points', methods=['POST'])
@admin_required
def update_fantasy_points() -> tuple[Response, int] | Response:
    """Update fantasy points for a player.

    Returns:
        JSON response with updated points.
    """
    data = request.get_json()
    if not data:
        return error_response('Request body is required')

    player_id = data.get('player_id')
    points = data.get('points', 0)

    # Validate player_id
    try:
        player_id = int(player_id)
    except (TypeError, ValueError):
        return error_response('Invalid player_id')

    # Validate points
    try:
        points = float(points)
    except (TypeError, ValueError):
        return error_response('Invalid points value')

    result = fantasy_service.update_player_points(player_id, points)
    return jsonify(result)


@api_bp.route('/fantasy/points/add', methods=['POST'])
@admin_required
def add_match_points() -> tuple[Response, int] | Response:
    """Add fantasy points for a specific match.

    Returns:
        JSON response with match points and total.
    """
    current_league = get_current_league()
    if not current_league:
        return error_response('No league selected')

    data = request.get_json()
    if not data:
        return error_response('Request body is required')

    player_id = data.get('player_id')
    match_number = data.get('match_number')
    points = data.get('points', 0)

    if not player_id or not match_number:
        return error_response('Player ID and match number required')

    # Validate player_id and match_number
    try:
        player_id = int(player_id)
        match_number = int(match_number)
    except (TypeError, ValueError):
        return error_response('Invalid player_id or match_number')

    if match_number <= 0:
        return error_response('Match number must be positive')

    # Validate points
    try:
        points = float(points)
    except (TypeError, ValueError):
        return error_response('Invalid points value')

    result = fantasy_service.add_match_points(
        player_id=player_id,
        match_number=match_number,
        points=points,
        league_id=current_league.id
    )
    return jsonify(result)


@api_bp.route('/fantasy/points/<int:player_id>', methods=['GET'])
def get_player_match_points(player_id: int) -> tuple[Response, int] | Response:
    """Get all match point entries for a player.

    Args:
        player_id: ID of the player.

    Returns:
        JSON response with player info and match entries.
    """
    current_league = get_current_league()
    league_id = current_league.id if current_league else None

    result = fantasy_service.get_player_match_points(player_id, league_id)
    return jsonify(result)


@api_bp.route('/fantasy/points/delete/<int:entry_id>', methods=['DELETE'])
@admin_required
def delete_match_points(entry_id: int) -> tuple[Response, int] | Response:
    """Delete a specific match point entry.

    Args:
        entry_id: ID of the entry to delete.

    Returns:
        JSON response with new total points.
    """
    current_league = get_current_league()
    if not current_league:
        return error_response('No league selected')

    result = fantasy_service.delete_match_points(entry_id, league_id=current_league.id)
    return jsonify(result)


# ==================== FANTASY AWARDS ====================

@api_bp.route('/fantasy/award', methods=['POST'])
@admin_required
def set_fantasy_award() -> tuple[Response, int] | Response:
    """Set a fantasy award (MVP, Orange Cap, Purple Cap).

    Returns:
        JSON response with award details.
    """
    current_league = get_current_league()
    if not current_league:
        return error_response('No league selected')

    data = request.get_json()
    if not data:
        return error_response('Request body is required')
    award_type = data.get('award_type')
    player_id = data.get('player_id')

    if player_id is not None:
        try:
            player_id = int(player_id)
        except (TypeError, ValueError):
            return error_response('Invalid player_id')

    result = fantasy_service.set_award(
        award_type=award_type,
        league_id=current_league.id,
        player_id=player_id
    )
    return jsonify(result)


@api_bp.route('/fantasy/awards', methods=['GET'])
def get_fantasy_awards() -> tuple[Response, int] | Response:
    """Get all fantasy awards for current league.

    Returns:
        JSON response with awards mapping.
    """
    current_league = get_current_league()
    if not current_league:
        return jsonify({'success': True, 'awards': {}})

    result = fantasy_service.get_awards(current_league.id)
    return jsonify(result)


@api_bp.route('/fantasy/players', methods=['GET'])
def get_fantasy_players() -> tuple[Response, int] | Response:
    """Get all sold players with fantasy points.

    Returns:
        JSON response with list of sold players.
    """
    current_league = get_current_league()
    if not current_league:
        return jsonify({'success': True, 'players': []})

    players = fantasy_service.get_sold_players(current_league.id)
    return jsonify({'success': True, 'players': players})


# ==================== DATA FETCHING ====================

@api_bp.route('/fantasy/fetch-awards', methods=['POST'])
@admin_required
def fetch_and_update_awards() -> tuple[Response, int] | Response:
    """Fetch Orange Cap, Purple Cap, and MVP from WPL and update awards.

    Returns:
        JSON response with fetch results.
    """
    current_league = get_current_league()
    if not current_league:
        return error_response('No league selected')

    result = fantasy_service.fetch_and_update_awards(current_league.id)
    return jsonify(result)


@api_bp.route('/fantasy/fetch-match-points', methods=['POST'])
@admin_required
def fetch_match_fantasy_points() -> tuple[Response, int] | Response:
    """Fetch all match scorecards and calculate fantasy points.

    Returns:
        JSON response with update summary.
    """
    current_league = get_current_league()
    if not current_league:
        return error_response('No league selected')

    result = fantasy_service.fetch_match_fantasy_points(current_league.id)
    return jsonify(result)


# ==================== EXCEL EXPORT ====================

# Position display mapping
_POSITION_CSV_MAP: dict[str, str] = {
    'Batter': 'BATSMAN',
    'Bowler': 'BOWLER',
    'Keeper': 'WK',
    'Allrounder': 'AR',
}

_POSITION_ORDER: list[str] = ['Batter', 'Bowler', 'Keeper', 'Allrounder']

# Colors for position labels (muted tones)
_POSITION_FILLS: dict[str, PatternFill] = {
    'Batter': PatternFill(start_color='C0504D', end_color='C0504D', fill_type='solid'),
    'Bowler': PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid'),
    'Keeper': PatternFill(start_color='E2A04F', end_color='E2A04F', fill_type='solid'),
    'Allrounder': PatternFill(start_color='548235', end_color='548235', fill_type='solid'),
}

# Team color mapping (name lowercase → hex fill color)
_TEAM_COLORS: dict[str, str] = {
    'csk': 'FFFF00',
    'chennai super kings': 'FFFF00',
    'mi': '004BA0',
    'mumbai indians': '004BA0',
    'kkr': '3A225D',
    'kolkata knight riders': '3A225D',
    'srh': 'FF6600',
    'sunrisers hyderabad': 'FF6600',
    'rcb': 'D32F2F',
    'royal challengers bengaluru': 'D32F2F',
    'royal challengers bangalore': 'D32F2F',
    'lsg': '00BFFF',
    'lucknow super giants': '00BFFF',
    'dc': '0078BC',
    'delhi capitals': '0078BC',
    'rr': 'E91E8C',
    'rajasthan royals': 'E91E8C',
    'gt': '1C1C2B',
    'gujarat titans': '1C1C2B',
    'pbks': 'ED1B24',
    'punjab kings': 'ED1B24',
}


def _get_team_fill(team_name: str) -> PatternFill:
    """Get fill color for a team name."""
    key = team_name.strip().lower()
    color = _TEAM_COLORS.get(key, '4472C4')
    return PatternFill(start_color=color, end_color=color, fill_type='solid')


def _team_font_color(team_name: str) -> str:
    """Return white or black font depending on team bg darkness."""
    light_teams = {'csk', 'chennai super kings', 'lsg', 'lucknow super giants'}
    return '000000' if team_name.strip().lower() in light_teams else 'FFFFFF'


@api_bp.route('/fantasy/export-csv', methods=['GET'])
def export_fantasy_csv() -> Response:
    """Export fantasy standings as a formatted Excel file.

    Returns:
        .xlsx file download response.
    """
    current_league = get_current_league()
    if not current_league:
        return error_response('No league selected')

    teams = Team.query.filter_by(
        league_id=current_league.id, is_deleted=False
    ).all()

    if not teams:
        return error_response('No teams found')

    # Build per-team data
    team_data: list[dict] = []
    for team in teams:
        sold = [p for p in team.players if p.status == 'sold' and not p.is_deleted]
        by_pos: dict[str, list] = {pos: [] for pos in _POSITION_ORDER}
        for p in sold:
            pos = p.position or 'Batter'
            if pos not in by_pos:
                pos = 'Batter'
            by_pos[pos].append(p)
        for pos in by_pos:
            by_pos[pos].sort(key=lambda x: x.fantasy_points or 0, reverse=True)
        team_data.append({
            'team': team,
            'players': sold,
            'by_pos': by_pos,
        })

    num_teams = len(team_data)
    total_cols = 1 + num_teams * 3

    wb = Workbook()
    ws = wb.active
    ws.title = 'Standings'

    # --- Shared styles ---
    thin = Side(style='thin')
    thick = Side(style='medium')
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    wrap_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    gray_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

    def team_border(ti: int) -> Border:
        """Border with thick left edge on first col of each team group."""
        left_side = thick if ti == 0 else thin
        return Border(left=left_side, right=thin, top=thin, bottom=thin)

    def team_border_right(ti: int) -> Border:
        """Border with thick right edge on last col of each team group."""
        return Border(left=thin, right=thick, top=thin, bottom=thin)

    def apply_team_cells(row_num: int, ti: int, col_start: int,
                         val1: object = '', val2: object = '', val3: object = '',
                         font: Font = None, fill: PatternFill = None,
                         align1: Alignment = None, align2: Alignment = None,
                         align3: Alignment = None) -> None:
        """Write 3 cells for a team group with proper borders."""
        for j, (val, al) in enumerate([
            (val1, align1 or center),
            (val2, align2 or center),
            (val3, align3 or center),
        ]):
            cell = ws.cell(row=row_num, column=col_start + j, value=val if val != '' else None)
            cell.alignment = al
            if font:
                cell.font = font
            if fill:
                cell.fill = fill
            # Thick left on PLAYER col, thick right on PRICE col
            l = thick if j == 0 else thin
            r = thick if j == 2 else thin
            cell.border = Border(left=l, right=r, top=thin, bottom=thin)

    row_num = 1  # current row pointer

    # ===== Row 1: blank =====
    row_num += 1

    # ===== Row 2: TEAM header =====
    ws.row_dimensions[row_num].height = 22
    team_label = ws.cell(row=row_num, column=1, value='TEAM')
    team_label.font = Font(bold=True, size=13)
    team_label.alignment = center
    team_label.border = Border(left=thick, right=thick, top=thick, bottom=thick)

    for ti, td in enumerate(team_data):
        col_start = 2 + ti * 3
        team_name = td['team'].name
        fill = _get_team_fill(team_name)
        fc = _team_font_color(team_name)
        team_font = Font(bold=True, color=fc, size=14)

        # Merge first, then apply styles to every cell in the range
        ws.merge_cells(start_row=row_num, start_column=col_start,
                       end_row=row_num, end_column=col_start + 2)
        for c in range(col_start, col_start + 3):
            mc = ws.cell(row=row_num, column=c)
            mc.fill = fill
            mc.font = team_font
            mc.alignment = center
            l = thick if c == col_start else thin
            r = thick if c == col_start + 2 else thin
            mc.border = Border(left=l, right=r, top=thick, bottom=thick)
        # Set value on the top-left cell
        ws.cell(row=row_num, column=col_start).value = _sanitize_excel_value(team_name)

    row_num += 1  # Row 3: blank
    row_num += 1  # Row 4: headers

    # ===== Row 4: PLAYER / POINTS / PRICE headers =====
    ws.cell(row=row_num, column=1).border = thin_border
    for ti in range(num_teams):
        col_start = 2 + ti * 3
        for j, hdr in enumerate(['PLAYER', 'POINTS', 'PRICE']):
            cell = ws.cell(row=row_num, column=col_start + j, value=hdr)
            cell.font = Font(bold=True, size=10)
            cell.alignment = center
            cell.fill = gray_fill
            l = thick if j == 0 else thin
            r = thick if j == 2 else thin
            cell.border = Border(left=l, right=r, top=thin, bottom=thick)

    row_num += 1

    # ===== Position group rows =====
    for pos in _POSITION_ORDER:
        csv_name = _POSITION_CSV_MAP[pos]
        pos_fill = _POSITION_FILLS[pos]
        team_players = [td['by_pos'][pos] for td in team_data]
        max_count = max((len(tp) for tp in team_players), default=0)

        for i in range(max_count):
            # Column A: position label on first row only
            if i == 0:
                cell = ws.cell(row=row_num, column=1, value=csv_name)
                cell.font = Font(bold=True, color='FFFFFF', size=11)
                cell.fill = pos_fill
                cell.alignment = center
                cell.border = Border(left=thick, right=thick, top=thin, bottom=thin)
            else:
                ws.cell(row=row_num, column=1).border = thin_border

            # Team columns
            for ti, tp in enumerate(team_players):
                col_start = 2 + ti * 3
                if i < len(tp):
                    p = tp[i]
                    overseas = '(F)' if p.country and p.country.lower() == 'overseas' else ''
                    price_cr = round((p.current_price or 0) / 10000000, 2)
                    pts = int(p.fantasy_points) if p.fantasy_points else 0
                    apply_team_cells(row_num, ti, col_start,
                                     _sanitize_excel_value(f'{p.name}{overseas}'), pts, price_cr)
                else:
                    apply_team_cells(row_num, ti, col_start)

            row_num += 1

    # ===== END row =====
    cell = ws.cell(row=row_num, column=1, value='END')
    cell.font = Font(bold=True, size=11)
    cell.alignment = center
    cell.border = Border(left=thick, right=thick, top=thick, bottom=thick)
    for ti in range(num_teams):
        col_start = 2 + ti * 3
        for j in range(3):
            c = ws.cell(row=row_num, column=col_start + j)
            l = thick if j == 0 else thin
            r = thick if j == 2 else thin
            c.border = Border(left=l, right=r, top=thick, bottom=thick)
    row_num += 1

    # ===== Summary rows =====
    # Only TOTAL POINTS, MONEY SPENT, MONEY REM get colored fills (matching reference)
    summary_defs = [
        ('TOTAL POINTS', 'pts',
         PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid'), 'FFFFFF',
         lambda td: int(sum(p.fantasy_points or 0 for p in td['players']))),
        ('MONEY SPENT', 'price',
         PatternFill(start_color='C0504D', end_color='C0504D', fill_type='solid'), 'FFFFFF',
         lambda td: round(((td['team'].initial_budget or 500000000) - td['team'].budget) / 10000000, 2)),
        ('MONEY REM', 'price',
         PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'), '000000',
         lambda td: round(td['team'].budget / 10000000, 2)),
        ('NUM PLAYERS', 'price',
         None, '000000',
         lambda td: len(td['players'])),
        ('\u2708\uFE0F PLAYERS', 'price',
         None, '000000',
         lambda td: sum(1 for p in td['players'] if p.country and p.country.lower() == 'overseas')),
    ]

    for label, col_type, fill, fc, value_fn in summary_defs:
        label_cell = ws.cell(row=row_num, column=1, value=label)
        label_cell.font = Font(bold=True, color=fc, size=11)
        if fill:
            label_cell.fill = fill
        label_cell.alignment = center
        label_cell.border = Border(left=thick, right=thick, top=thin, bottom=thin)

        sfont = Font(bold=True, color=fc, size=11)
        for ti, td in enumerate(team_data):
            col_start = 2 + ti * 3
            val = value_fn(td)
            idx = 1 if col_type == 'pts' else 2
            vals = [None, None, None]
            vals[idx] = val
            for j in range(3):
                cell = ws.cell(row=row_num, column=col_start + j, value=vals[j])
                cell.alignment = center
                # Only color the cell that has the value, not empty ones
                if vals[j] is not None:
                    cell.font = sfont
                    if fill:
                        cell.fill = fill
                l = thick if j == 0 else thin
                r = thick if j == 2 else thin
                cell.border = Border(left=l, right=r, top=thin, bottom=thin)

        row_num += 1

    # ===== Blank row =====
    row_num += 1

    # ===== Awards =====
    award_configs = [
        ('MVP', 'mvp',
         PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid'), '000000'),
        ('Orange Cap', 'orange_cap',
         PatternFill(start_color='F4B183', end_color='F4B183', fill_type='solid'), '000000'),
        ('Purple Cap', 'purple_cap',
         PatternFill(start_color='B4A7D6', end_color='B4A7D6', fill_type='solid'), '000000'),
    ]

    # Place awards in columns H and I (8 and 9)
    award_label_col = 8
    award_name_col = 9

    for label, key, fill, fc in award_configs:
        award = FantasyAward.query.filter_by(
            award_type=key, league_id=current_league.id
        ).first()

        lc = ws.cell(row=row_num, column=award_label_col, value=label)
        lc.font = Font(bold=True, color=fc, size=11)
        lc.fill = fill
        lc.alignment = center
        lc.border = thin_border

        nc = ws.cell(row=row_num, column=award_name_col)
        nc.border = thin_border
        nc.alignment = left_align
        if award and award.player:
            nc.value = _sanitize_excel_value(award.player.name)
            nc.font = Font(bold=True, size=11)

        row_num += 1

    # Note about amounts — bottom right area
    row_num += 1
    note_col = max(total_cols - 1, award_name_col + 2)
    note_cell = ws.cell(row=row_num, column=note_col, value='*ALL amount in CR*')
    note_cell.font = Font(bold=True, italic=True, size=10)

    # ===== Column widths =====
    ws.column_dimensions['A'].width = 14
    for ti in range(num_teams):
        col_start = 2 + ti * 3
        ws.column_dimensions[get_column_letter(col_start)].width = 18
        ws.column_dimensions[get_column_letter(col_start + 1)].width = 7
        ws.column_dimensions[get_column_letter(col_start + 2)].width = 6

    # Freeze panes so headers stay visible when scrolling
    ws.freeze_panes = 'B5'

    # Write to buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    league_name = current_league.name.replace(' ', '_')
    filename = f'{league_name}_standings.xlsx'

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename,
    )
